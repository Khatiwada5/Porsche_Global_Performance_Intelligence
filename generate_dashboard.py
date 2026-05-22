from __future__ import annotations

import math
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
import plotly.graph_objects as go
import plotly.io as pio
from plotly.offline.offline import get_plotlyjs
from plotly.subplots import make_subplots


ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
CSV_PATH = DATA_DIR / "porsche_global_performance_synthetic.csv"
EXCEL_PATH = DATA_DIR / "Porsche_Performance_Data.xlsx"
HTML_PATH = ROOT / "index.html"

PORSCHE_RED = "#D5001C"
PORSCHE_GOLD = "#C8A96E"
BG = "#0A0A0F"
CARD = "#14141C"
TEXT = "#FFFFFF"
MUTED = "#B8B8C8"
GRID = "rgba(255,255,255,0.08)"

REPO_URL = "https://github.com/Khatiwada5/Porsche_Global_Performance_Intelligence"
LIVE_DATA_URL = f"{REPO_URL}/raw/main/data/Porsche_Performance_Data.xlsx"
SOURCE_URL = f"{REPO_URL}/blob/main/generate_dashboard.py"


def sports_car_svg(class_name: str = "sports-car", width: int = 60, height: int = 28) -> str:
    return f"""<svg class="{class_name}" width="{width}" height="{height}" viewBox="0 0 120 56" aria-hidden="true" focusable="false" xmlns="http://www.w3.org/2000/svg">
        <path fill="currentColor" d="M9 36.6c6.9-7.8 15.8-12.4 29.6-13.7 7.9-8.7 18.7-13.5 31.5-13.5 10.2 0 18.7 3.1 27.2 10.3 7.6.8 13.2 3 17.9 7.4 1.5 1.4 1 3.9-.9 4.5l-9.9 3.2c-1.4.5-2.9-.1-3.7-1.4-2.3-3.8-6.4-6.1-11.1-6.1-5.9 0-10.8 3.8-12.5 9.1H44.2c-1.7-5.3-6.7-9.1-12.6-9.1-5.7 0-10.6 3.6-12.4 8.7-.5 1.5-2 2.4-3.5 2.1l-5.4-1c-1.1-.2-1.7-1.5-.9-2.5Z"/>
        <path fill="currentColor" d="M43.9 21.9c7-6 15.8-9.1 26.1-9.1 7.9 0 14.5 2.1 20.4 6.6l-14.7 1.7c-12.8 1.5-22.7 1.8-31.8.8Z" opacity="0.72"/>
        <path fill="currentColor" d="M31.6 31.1a8.8 8.8 0 1 0 0 17.6 8.8 8.8 0 0 0 0-17.6Zm58 0a8.8 8.8 0 1 0 0 17.6 8.8 8.8 0 0 0 0-17.6Z"/>
        <path fill="currentColor" d="M31.6 35.7a4.2 4.2 0 1 0 0 8.4 4.2 4.2 0 0 0 0-8.4Zm58 0a4.2 4.2 0 1 0 0 8.4 4.2 4.2 0 0 0 0-8.4Z" fill-opacity="0.28"/>
        <path fill="none" stroke="currentColor" stroke-width="3.2" stroke-linecap="round" d="M16.8 34.2c13.6-6.5 26.2-7.6 41.7-6.3 13.9 1.1 28.5-1.3 42.5-3.2"/>
    </svg>"""


def weighted_alloc(total: int, weights: dict[str, float]) -> dict[str, int]:
    raw = {key: total * weight for key, weight in weights.items()}
    floors = {key: int(math.floor(value)) for key, value in raw.items()}
    remainder = total - sum(floors.values())
    ranked = sorted(raw, key=lambda key: raw[key] - floors[key], reverse=True)
    for key in ranked[:remainder]:
        floors[key] += 1
    return floors


def build_dataset() -> pd.DataFrame:
    # Anchored to Porsche's public worldwide delivery totals from annual reports/newsroom.
    annual_totals = {
        2018: 256_255,
        2019: 280_800,
        2020: 272_162,
        2021: 301_915,
        2022: 309_884,
        2023: 320_221,
        2024: 310_718,
    }

    reported_model_mix = {
        2018: {"911": 35_573, "718": 24_750, "Macan": 86_031, "Cayenne": 71_458, "Panamera": 38_443, "Taycan": 0},
        2019: {"911": 34_800, "718": 20_467, "Macan": 99_944, "Cayenne": 92_055, "Panamera": 32_562, "Taycan": 972},
        2020: {"911": 34_328, "718": 21_784, "Macan": 78_124, "Cayenne": 92_860, "Panamera": 24_307, "Taycan": 20_759},
        2021: {"911": 38_464, "718": 20_502, "Macan": 88_362, "Cayenne": 83_071, "Panamera": 30_220, "Taycan": 41_296},
        2022: {"911": 40_410, "718": 18_203, "Macan": 86_724, "Cayenne": 95_604, "Panamera": 34_142, "Taycan": 34_801},
        2023: {"911": 50_146, "718": 20_518, "Macan": 87_355, "Cayenne": 87_553, "Panamera": 34_020, "Taycan": 40_629},
        2024: {"911": 50_941, "718": 23_670, "Macan": 82_795, "Cayenne": 102_889, "Panamera": 29_587, "Taycan": 20_836},
    }

    region_weights = {
        2018: {"North America": 0.29, "Europe": 0.28, "China": 0.31, "Asia Pacific": 0.08, "Rest of World": 0.04},
        2019: {"North America": 0.29, "Europe": 0.27, "China": 0.31, "Asia Pacific": 0.09, "Rest of World": 0.04},
        2020: {"North America": 0.26, "Europe": 0.29, "China": 0.33, "Asia Pacific": 0.08, "Rest of World": 0.04},
        2021: {"North America": 0.28, "Europe": 0.29, "China": 0.32, "Asia Pacific": 0.07, "Rest of World": 0.04},
        2022: {"North America": 0.26, "Europe": 0.30, "China": 0.30, "Asia Pacific": 0.09, "Rest of World": 0.05},
        2023: {"North America": 0.27, "Europe": 0.31, "China": 0.25, "Asia Pacific": 0.11, "Rest of World": 0.06},
        2024: {"North America": 0.2786, "Europe": 0.3000, "China": 0.1830, "Asia Pacific": 0.1450, "Rest of World": 0.0934},
    }

    region_price_index = {
        "North America": 1.08,
        "Europe": 1.03,
        "China": 1.14,
        "Asia Pacific": 1.01,
        "Rest of World": 0.96,
    }

    model_base_price = {
        "911": 154_000,
        "Cayenne": 112_000,
        "Macan": 78_000,
        "Panamera": 128_000,
        "Taycan": 121_000,
        "718": 84_000,
        "Cayenne Coupe": 128_000,
    }

    model_satisfaction = {
        "911": 9.3,
        "Cayenne": 8.8,
        "Macan": 8.6,
        "Panamera": 8.7,
        "Taycan": 8.9,
        "718": 8.8,
        "Cayenne Coupe": 8.9,
    }

    powertrain_mix = {
        "911": {"ICE": 1.00},
        "718": {"ICE": 1.00},
        "Macan": {"ICE": 0.92, "Electric": 0.08},
        "Cayenne": {"ICE": 0.58, "Hybrid": 0.42},
        "Cayenne Coupe": {"ICE": 0.55, "Hybrid": 0.45},
        "Panamera": {"ICE": 0.38, "Hybrid": 0.62},
        "Taycan": {"Electric": 1.00},
    }

    rows = []
    for year, total in annual_totals.items():
        reported = reported_model_mix[year].copy()
        cayenne_coupe_share = 0.00 if year < 2020 else {2020: 0.10, 2021: 0.13, 2022: 0.16, 2023: 0.18, 2024: 0.19}[year]
        cayenne_coupe_units = round(reported["Cayenne"] * cayenne_coupe_share)
        reported["Cayenne"] -= cayenne_coupe_units
        reported["Cayenne Coupe"] = cayenne_coupe_units

        if sum(reported.values()) != total:
            reported["Macan"] += total - sum(reported.values())

        regional_units = weighted_alloc(total, region_weights[year])
        model_weights = {model: units / total for model, units in reported.items() if units > 0}

        for region, region_total in regional_units.items():
            model_alloc = weighted_alloc(region_total, model_weights)
            for model, units in model_alloc.items():
                if units == 0:
                    continue
                for powertrain, share in powertrain_mix[model].items():
                    if model == "Macan" and powertrain == "Electric" and year < 2024:
                        continue
                    effective_share = share
                    if model == "Macan":
                        if year < 2024:
                            effective_share = 1.00
                        else:
                            effective_share = 0.12 if powertrain == "Electric" else 0.88
                    split_units = round(units * effective_share)
                    if split_units == 0:
                        continue
                    maturity = 1 + 0.021 * (year - 2018)
                    mix_lift = {"ICE": 1.00, "Hybrid": 1.11, "Electric": 1.18}[powertrain]
                    avg_price = model_base_price[model] * region_price_index[region] * maturity * mix_lift
                    satisfaction = model_satisfaction[model]
                    satisfaction += {"North America": 0.05, "Europe": 0.03, "China": -0.15, "Asia Pacific": 0.02, "Rest of World": -0.03}[region]
                    satisfaction += {"ICE": 0.00, "Hybrid": 0.05, "Electric": 0.07}[powertrain]
                    satisfaction += 0.03 * (year - 2021)
                    rows.append(
                        {
                            "Year": year,
                            "Model": model,
                            "Region": region,
                            "Units Sold": split_units,
                            "Revenue in EUR millions": round(split_units * avg_price / 1_000_000, 2),
                            "Average Transaction Price in EUR": round(avg_price, 0),
                            "Customer Satisfaction Score (1-10)": round(min(max(satisfaction, 7.6), 9.7), 1),
                            "Dealer Count": {
                                "North America": 205,
                                "Europe": 310,
                                "China": 155,
                                "Asia Pacific": 118,
                                "Rest of World": 86,
                            }[region]
                            + int((year - 2018) * {"North America": 2.0, "Europe": 1.0, "China": 1.5, "Asia Pacific": 2.5, "Rest of World": 1.3}[region]),
                            "Electric vs ICE vs Hybrid classification": powertrain,
                        }
                    )

    df = pd.DataFrame(rows)
    for year, target in annual_totals.items():
        actual = int(df.loc[df["Year"] == year, "Units Sold"].sum())
        difference = target - actual
        if difference:
            idx = df[df["Year"] == year]["Units Sold"].idxmax()
            df.loc[idx, "Units Sold"] += difference
            df.loc[idx, "Revenue in EUR millions"] = round(
                df.loc[idx, "Units Sold"] * df.loc[idx, "Average Transaction Price in EUR"] / 1_000_000,
                2,
            )
    return df


def chart_layout(fig: go.Figure, title: str, subtitle: str, height: int = 520) -> go.Figure:
    fig.update_layout(
        title={"text": f"<b>{title}</b><br><span style='font-size:13px;color:{MUTED}'>{subtitle}</span>", "x": 0.02},
        paper_bgcolor=CARD,
        plot_bgcolor=CARD,
        font={"family": "Inter, Arial, sans-serif", "color": TEXT},
        margin={"l": 70, "r": 40, "t": 92, "b": 60},
        height=height,
        legend={"orientation": "h", "yanchor": "bottom", "y": -0.22, "x": 0},
        colorway=[PORSCHE_RED, PORSCHE_GOLD, "#FFFFFF", "#8A8D98", "#E6C98C", "#7A0010", "#D8D8E2"],
    )
    fig.update_xaxes(gridcolor=GRID, zerolinecolor=GRID)
    fig.update_yaxes(gridcolor=GRID, zerolinecolor=GRID)
    return fig


def plot_html(fig: go.Figure) -> str:
    return pio.to_html(fig, include_plotlyjs=False, full_html=False, config={"displayModeBar": False, "responsive": True})


def build_forecast_frame(df: pd.DataFrame) -> pd.DataFrame:
    yearly_total = df.groupby("Year", as_index=False).agg(
        actual_units=("Units Sold", "sum"),
        actual_revenue_eur_millions=("Revenue in EUR millions", "sum"),
    )
    years = yearly_total["Year"].to_numpy()
    units = yearly_total["actual_units"].to_numpy()
    coef = np.polyfit(years, units, deg=1)
    forecast_years = np.arange(2018, 2028)
    baseline = np.polyval(coef, forecast_years)
    residual_std = float(np.std(units - np.polyval(coef, years)))

    forecast_df = pd.DataFrame(
        {
            "Year": forecast_years,
            "Actual Units": [int(yearly_total.loc[yearly_total["Year"] == year, "actual_units"].iloc[0]) if year in set(years) else np.nan for year in forecast_years],
            "Regression Baseline Units": np.round(baseline, 0),
            "Confidence Lower Units": np.round(baseline - residual_std * 1.65, 0),
            "Confidence Upper Units": np.round(baseline + residual_std * 1.65, 0),
            "Conservative Scenario Units": np.nan,
            "Optimistic Scenario Units": np.nan,
        }
    )
    future_years = np.arange(2025, 2028)
    conservative = np.polyval(coef, future_years) * np.array([0.985, 0.995, 1.005])
    optimistic = np.polyval(coef, future_years) * np.array([1.025, 1.055, 1.085])
    forecast_df.loc[forecast_df["Year"].isin(future_years), "Conservative Scenario Units"] = np.round(conservative, 0)
    forecast_df.loc[forecast_df["Year"].isin(future_years), "Optimistic Scenario Units"] = np.round(optimistic, 0)
    return forecast_df


def export_excel(df: pd.DataFrame, path: Path) -> None:
    regional = df.groupby(["Year", "Region"], as_index=False).agg(
        Units_Sold=("Units Sold", "sum"),
        Revenue_EUR_Millions=("Revenue in EUR millions", "sum"),
    )
    powertrain = df.groupby(["Year", "Model", "Electric vs ICE vs Hybrid classification"], as_index=False).agg(
        Units_Sold=("Units Sold", "sum"),
        Revenue_EUR_Millions=("Revenue in EUR millions", "sum"),
    )
    forecast_df = build_forecast_frame(df)

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Global Sales", index=False)
        regional.to_excel(writer, sheet_name="Regional Breakdown", index=False)
        powertrain.to_excel(writer, sheet_name="Powertrain Mix", index=False)
        forecast_df.to_excel(writer, sheet_name="Forecasting Data", index=False)

        workbook = writer.book
        for worksheet in workbook.worksheets:
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(fill_type="solid", fgColor="D5001C")
            stripe_fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
                if row_idx % 2 == 0:
                    for cell in row:
                        cell.fill = stripe_fill
            for column in worksheet.columns:
                max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
                worksheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 42)
            worksheet.freeze_panes = "A2"


def make_dashboard(df: pd.DataFrame) -> str:
    yearly_model = df.groupby(["Year", "Model"], as_index=False)["Units Sold"].sum()
    yearly_total = df.groupby("Year", as_index=False).agg(
        units=("Units Sold", "sum"),
        revenue=("Revenue in EUR millions", "sum"),
    )
    y2024 = df[df["Year"] == 2024]
    y2023 = df[df["Year"] == 2023]
    revenue_2024 = y2024["Revenue in EUR millions"].sum()
    revenue_2023 = y2023["Revenue in EUR millions"].sum()
    yoy_growth = (revenue_2024 / revenue_2023 - 1) * 100
    best_model = y2024.groupby("Model")["Units Sold"].sum().idxmax()
    region_growth = (
        df[df["Year"].isin([2023, 2024])]
        .groupby(["Year", "Region"])["Revenue in EUR millions"]
        .sum()
        .unstack(0)
    )
    region_growth["growth"] = region_growth[2024] / region_growth[2023] - 1
    fastest_region = region_growth["growth"].idxmax()

    trend = go.Figure()
    for model in ["911", "Cayenne", "Macan", "Panamera", "Taycan", "718", "Cayenne Coupe"]:
        data = yearly_model[yearly_model["Model"] == model]
        trend.add_trace(go.Scatter(x=data["Year"], y=data["Units Sold"], mode="lines+markers", name=model, line={"width": 3}))
    trend.add_vline(x=2019, line_color=PORSCHE_RED, line_dash="dash")
    trend.add_annotation(x=2019, y=yearly_model["Units Sold"].max() * 1.02, text="Taycan launch", showarrow=True, arrowcolor=PORSCHE_RED, font={"color": PORSCHE_RED})
    chart_layout(trend, "Sales Trend Analysis", "Global deliveries by model, 2018-2024", 560)

    region_rev = y2024.groupby("Region", as_index=False)["Revenue in EUR millions"].sum().sort_values("Revenue in EUR millions")
    bars = go.Figure(go.Bar(x=region_rev["Revenue in EUR millions"], y=region_rev["Region"], orientation="h", marker={"color": [PORSCHE_RED if r == fastest_region else PORSCHE_GOLD for r in region_rev["Region"]]}))
    chart_layout(bars, "Regional Revenue Contribution", "2024 revenue in EUR millions by region", 470)
    bars.update_xaxes(title="Revenue (EUR millions)")

    market = y2024.groupby("Region", as_index=False)["Units Sold"].sum()
    donut = go.Figure(go.Pie(labels=market["Region"], values=market["Units Sold"], hole=0.62, marker={"colors": [PORSCHE_RED, PORSCHE_GOLD, "#F4F4F8", "#777A86", "#3B3B48"]}))
    chart_layout(donut, "Regional Market Share", "2024 delivery mix by region", 470)

    model_profit = y2024.groupby("Model", as_index=False).agg(
        units=("Units Sold", "sum"),
        revenue=("Revenue in EUR millions", "sum"),
        atp=("Average Transaction Price in EUR", "mean"),
    )
    scatter = go.Figure(
        go.Scatter(
            x=model_profit["units"],
            y=model_profit["atp"],
            mode="markers+text",
            text=model_profit["Model"],
            textposition="top center",
            marker={
                "size": model_profit["revenue"] / model_profit["revenue"].max() * 72 + 18,
                "color": model_profit["revenue"],
                "colorscale": [[0, "#3B3B48"], [0.55, PORSCHE_GOLD], [1, PORSCHE_RED]],
                "showscale": True,
                "colorbar": {"title": "Revenue"},
                "line": {"color": "white", "width": 1},
            },
        )
    )
    chart_layout(scatter, "Model Profitability Matrix", "Volume versus average transaction price, sized by revenue", 560)
    scatter.update_xaxes(title="Units Sold")
    scatter.update_yaxes(title="Average Transaction Price (EUR)")

    powertrain = df.groupby(["Year", "Electric vs ICE vs Hybrid classification"], as_index=False).agg(
        units=("Units Sold", "sum"),
        revenue=("Revenue in EUR millions", "sum"),
    )
    powertrain["revenue_per_unit"] = powertrain["revenue"] * 1_000_000 / powertrain["units"]
    transition = make_subplots(specs=[[{"secondary_y": True}]])
    for pt, color in [("ICE", "#F4F4F8"), ("Hybrid", PORSCHE_GOLD), ("Electric", PORSCHE_RED)]:
        data = powertrain[powertrain["Electric vs ICE vs Hybrid classification"] == pt]
        transition.add_trace(go.Bar(x=data["Year"], y=data["units"], name=f"{pt} units", marker_color=color), secondary_y=False)
        transition.add_trace(go.Scatter(x=data["Year"], y=data["revenue_per_unit"], mode="lines+markers", name=f"{pt} EUR/unit", line={"dash": "dot", "width": 2, "color": color}), secondary_y=True)
    transition.update_layout(barmode="stack")
    chart_layout(transition, "Electric Transition Tracker", "Powertrain unit mix with revenue-per-unit overlay", 600)
    transition.update_yaxes(title="Units Sold", secondary_y=False)
    transition.update_yaxes(title="Revenue per Unit (EUR)", secondary_y=True)

    years = yearly_total["Year"].to_numpy()
    units = yearly_total["units"].to_numpy()
    coef = np.polyfit(years, units, deg=1)
    forecast_years = np.arange(2018, 2028)
    baseline = np.polyval(coef, forecast_years)
    residual_std = float(np.std(units - np.polyval(coef, years)))
    future_years = np.arange(2025, 2028)
    conservative = np.polyval(coef, future_years) * np.array([0.985, 0.995, 1.005])
    optimistic = np.polyval(coef, future_years) * np.array([1.025, 1.055, 1.085])
    forecast = go.Figure()
    forecast.add_trace(go.Scatter(x=years, y=units, mode="lines+markers", name="Actual deliveries", line={"color": PORSCHE_RED, "width": 4}))
    forecast.add_trace(go.Scatter(x=future_years, y=conservative, mode="lines+markers", name="Conservative scenario", line={"color": PORSCHE_GOLD, "width": 3}))
    forecast.add_trace(go.Scatter(x=future_years, y=optimistic, mode="lines+markers", name="Optimistic EV adoption", line={"color": "#FFFFFF", "width": 3}))
    upper = np.polyval(coef, forecast_years) + residual_std * 1.65
    lower = np.polyval(coef, forecast_years) - residual_std * 1.65
    forecast.add_trace(go.Scatter(x=forecast_years, y=upper, mode="lines", line={"width": 0}, showlegend=False))
    forecast.add_trace(go.Scatter(x=forecast_years, y=lower, mode="lines", fill="tonexty", fillcolor="rgba(213,0,28,0.16)", line={"width": 0}, name="Confidence interval"))
    chart_layout(forecast, "Forecasting", "Linear regression projection, 2025-2027, with scenario adjustment", 560)
    forecast.update_xaxes(title="Year", dtick=1)
    forecast.update_yaxes(title="Global Deliveries")

    insights = {
        "trend": "Porsche's delivery story is now a portfolio balancing act. SUVs keep the company resilient at scale, while the 911 has become a stronger premium anchor and Taycan added an electrified halo without fully replacing combustion demand.",
        "regional": "North America and Europe now provide the most dependable revenue base, while China softened materially in 2024. Porsche should defend pricing in mature regions and rebuild China momentum with localized EV, software, and dealer-experience investments.",
        "profit": "Cayenne and Macan remain the volume engines, but the 911 protects the brand's pricing ceiling. Low-volume GT3/RS-style halo derivatives are not modeled as separate rows, yet the 911 pricing premium reflects how those variants defend Porsche's margin architecture.",
        "transition": "Electrification is helping average revenue per unit, especially where Taycan and hybrid Cayenne/Panamera trims lift transaction prices. The risk is not margin dilution; it is adoption volatility during model changeovers and uneven regional charging readiness.",
        "forecast": "The conservative path assumes EV launches restore mix slowly after the 2024 transition year. The optimistic case requires Macan Electric scale, Taycan recovery, and disciplined supply allocation to high-ATP regions.",
    }

    kpis = [
        ("01", "Total Revenue 2024", f"EUR {revenue_2024:,.0f}M", "Synthetic estimate from unit mix and ATP", {"target": round(revenue_2024), "prefix": "EUR ", "suffix": "M", "decimals": 0, "signed": "false"}),
        ("02", "YoY Growth", f"{yoy_growth:+.1f}%", "Revenue growth versus 2023", {"target": round(yoy_growth, 1), "prefix": "", "suffix": "%", "decimals": 1, "signed": "true"}),
        ("03", "Best Selling Model", best_model, "2024 global deliveries", None),
        ("04", "Fastest Growing Region", fastest_region, f"{region_growth.loc[fastest_region, 'growth'] * 100:+.1f}% revenue YoY", None),
    ]
    kpi_cards = []
    for marker, label, value, sub, count_meta in kpis:
        attrs = ""
        if count_meta:
            attrs = (
                f" data-count-target='{count_meta['target']}'"
                f" data-count-prefix='{count_meta['prefix']}'"
                f" data-count-suffix='{count_meta['suffix']}'"
                f" data-count-decimals='{count_meta['decimals']}'"
                f" data-count-signed='{count_meta['signed']}'"
            )
        kpi_cards.append(f"<article class='kpi-card'><div class='kpi-marker'>{marker}</div><p>{label}</p><strong{attrs}>{value}</strong><span>{sub}</span></article>")
    kpi_html = "\n".join(kpi_cards)

    recommendations = [
        ("Prioritize Electrified SUV Scale", "Macan Electric and Cayenne Hybrid are the clearest bridge between Porsche's volume base and its EV transition. Protect allocation for North America and Europe where revenue per unit is strongest.", "Higher EV penetration without abandoning the models that fund operating leverage."),
        ("Rebuild China With Product-Led Local Relevance", "China remains strategically important but no longer behaves like an automatic growth market. Porsche should pair localized digital services with selective EV derivatives and sharper dealer experience management.", "Stabilized regional share and improved customer satisfaction in the most volatile large market."),
        ("Keep the 911 as the Pricing North Star", "The 911 is not the highest-volume model, but it has become the strongest signal of brand scarcity and margin resilience. Continue using GT, RS, and special-edition logic to preserve transaction-price discipline across the range.", "Sustained premium perception and pricing power even as the portfolio electrifies."),
        ("Manage Transition Years Like Portfolio Investments", "2024 shows how model changeovers can obscure underlying demand. Porsche should track launch readiness, regional mix, and ATP by powertrain in one operating cadence rather than treating EV adoption as a separate program.", "Cleaner capital allocation, fewer supply-demand mismatches, and better executive visibility."),
    ]

    recommendation_html = "\n".join(
        f"<article class='recommendation'><div class='rec-number'>{i:02d}</div><div><h3>{title}</h3><p>{rationale}</p><span class='impact-badge'>Expected Impact</span><strong>{impact}</strong></div></article>"
        for i, (title, rationale, impact) in enumerate(recommendations, start=1)
    )

    data_dictionary = [
        ("Year", "Integer", "Reporting year from 2018 through 2024."),
        ("Model", "Text", "Porsche model line used for portfolio analysis."),
        ("Region", "Text", "Primary sales region used for delivery and revenue allocation."),
        ("Units Sold", "Integer", "Synthetic deliveries anchored to public annual Porsche totals."),
        ("Revenue in EUR millions", "Float", "Estimated revenue from units sold and modeled transaction price."),
        ("Average Transaction Price in EUR", "Float", "Modeled average transaction price by model, region, year, and powertrain."),
        ("Customer Satisfaction Score (1-10)", "Float", "Synthetic satisfaction score reflecting model and regional assumptions."),
        ("Dealer Count", "Integer", "Estimated dealer count by region and year."),
        ("Electric vs ICE vs Hybrid classification", "Text", "Powertrain category used in the electric transition analysis."),
    ]
    dictionary_rows = "\n".join(
        f"<tr><td>{column}</td><td>{dtype}</td><td>{description}</td></tr>"
        for column, dtype, description in data_dictionary
    )

    plotly_bundle = f"<script>{get_plotlyjs()}</script>"
    header_car = sports_car_svg("header-car", 60, 28)
    section_car = sports_car_svg("section-car", 120, 56)
    footer_car = sports_car_svg("footer-car", 92, 42)

    return f"""<!doctype html>
<html lang="en">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>Porsche Global Performance Intelligence</title>
  {plotly_bundle}
  <style>
    :root {{
      --bg: {BG};
      --card: {CARD};
      --red: {PORSCHE_RED};
      --gold: {PORSCHE_GOLD};
      --text: {TEXT};
      --muted: {MUTED};
      --line: rgba(255,255,255,0.12);
      --green: #22C55E;
    }}
    * {{ box-sizing: border-box; }}
    html {{ scroll-behavior: smooth; }}
    body {{
      margin: 0;
      background: var(--bg);
      color: var(--text);
      font-family: Inter, Arial, Helvetica, sans-serif;
      letter-spacing: 0;
    }}
    header {{
      min-height: 72vh;
      display: grid;
      align-items: center;
      border-bottom: 1px solid var(--line);
      background:
        linear-gradient(120deg, rgba(213,0,28,0.22), transparent 34%),
        linear-gradient(170deg, rgba(200,169,110,0.18), transparent 42%),
        var(--bg);
    }}
    .wrap {{ width: min(1180px, calc(100% - 36px)); margin: 0 auto; }}
    .brandline {{ display: inline-flex; align-items: center; gap: 16px; margin-bottom: 42px; color: var(--red); }}
    .header-car {{ flex: 0 0 auto; filter: drop-shadow(0 0 14px rgba(213,0,28,0.42)); }}
    .wordmark {{ font-weight: 900; letter-spacing: 0.18em; color: var(--text); font-size: 14px; }}
    .wordmark::after {{ content: ""; display: block; width: 96px; height: 3px; margin-top: 14px; background: var(--red); box-shadow: 0 0 22px rgba(213,0,28,0.55); }}
    h1 {{ font-size: clamp(42px, 8vw, 92px); line-height: 0.94; margin: 0; max-width: 980px; letter-spacing: 0; }}
    .hero-copy {{ color: var(--muted); max-width: 760px; font-size: 18px; line-height: 1.65; margin: 26px 0 0; }}
    nav {{
      position: sticky;
      top: 0;
      z-index: 20;
      backdrop-filter: blur(16px);
      background: rgba(10,10,15,0.88);
      border-bottom: 1px solid rgba(213,0,28,0.38);
      overflow-x: auto;
    }}
    nav .wrap {{ display: flex; gap: 8px; padding: 12px 0; }}
    nav a {{
      position: relative;
      color: var(--muted);
      text-decoration: none;
      white-space: nowrap;
      padding: 10px 12px;
      border-bottom: 2px solid transparent;
      font-size: 13px;
      font-weight: 700;
      transition: color 180ms ease, border-color 180ms ease, background 180ms ease;
    }}
    nav a::after {{ content: ""; position: absolute; left: 12px; right: 12px; bottom: 4px; height: 2px; background: var(--red); transform: scaleX(0); transform-origin: center; transition: transform 220ms ease; }}
    nav a:hover, nav a.active {{ color: var(--red); border-color: transparent; background: rgba(213,0,28,0.10); }}
    nav a:hover::after, nav a.active::after {{ transform: scaleX(1); }}
    main {{ padding: 34px 0 72px; }}
    section {{ position: relative; scroll-margin-top: 74px; margin: 34px 0 56px; opacity: 0; transform: translateY(30px); transition: opacity 700ms ease, transform 700ms ease; }}
    section.visible {{ opacity: 1; transform: translateY(0); }}
    section > *:not(.section-watermark) {{ position: relative; z-index: 1; }}
    .section-watermark {{ position: absolute; top: 8px; right: 4px; z-index: 0; color: var(--red); opacity: 0.04; pointer-events: none; }}
    .section-title {{ display: flex; justify-content: space-between; gap: 18px; align-items: end; margin-bottom: 18px; }}
    h2 {{ font-size: 28px; margin: 0; color: var(--text); position: relative; padding-left: 16px; }}
    h2::before {{ content: ""; position: absolute; left: 0; top: 6px; bottom: 6px; width: 4px; border-radius: 999px; background: var(--red); box-shadow: 0 0 18px rgba(213,0,28,0.62); }}
    .section-title p {{ color: var(--muted); margin: 0; max-width: 560px; line-height: 1.5; }}
    .kpi-grid {{ display: grid; grid-template-columns: repeat(4, 1fr); gap: 18px; }}
    .kpi-card, .chart-card, .recommendation, .data-card {{
      background: var(--card);
      border: 1px solid var(--line);
      box-shadow: 0 18px 50px rgba(0,0,0,0.32);
      border-radius: 8px;
    }}
    .kpi-card {{ position: relative; padding: 30px; border: 1px solid rgba(213,0,28,0.34); border-top: 4px solid var(--red); min-height: 214px; overflow: hidden; transition: transform 240ms ease, border-color 240ms ease, box-shadow 240ms ease; }}
    .kpi-card::after {{ content: ""; position: absolute; inset: auto -32px -44px auto; width: 120px; height: 120px; border-radius: 999px; background: rgba(213,0,28,0.16); filter: blur(14px); }}
    .kpi-card:hover {{ transform: translateY(-6px); border-color: var(--red); box-shadow: 0 0 20px rgba(213,0,28,0.15), 0 26px 72px rgba(213,0,28,0.24), 0 18px 50px rgba(0,0,0,0.38); }}
    .kpi-marker {{ color: var(--gold); font-size: 12px; font-weight: 900; letter-spacing: 0.12em; margin-bottom: 28px; }}
    .kpi-card p {{ margin: 0 0 18px; color: var(--muted); font-size: 13px; text-transform: uppercase; font-weight: 900; }}
    .kpi-card strong {{ display: block; color: var(--red); font-size: clamp(34px, 3.1vw, 52px); line-height: 1; margin-bottom: 16px; text-shadow: 0 0 26px rgba(213,0,28,0.28); }}
    .kpi-card span {{ color: var(--gold); font-size: 13px; line-height: 1.45; }}
    .chart-card {{ padding: 10px; margin-bottom: 18px; border-color: rgba(213,0,28,0.2); border-left: 3px solid var(--red); }}
    .chart-grid {{ display: grid; grid-template-columns: 1fr 1fr; gap: 18px; }}
    details {{
      background: rgba(255,255,255,0.04);
      border: 1px solid var(--line);
      border-left: 3px solid var(--red);
      border-radius: 8px;
      padding: 14px 16px;
      margin-top: 12px;
    }}
    summary {{ cursor: pointer; font-weight: 800; color: var(--red); }}
    details p {{ color: var(--muted); line-height: 1.65; margin: 12px 0 0; }}
    .data-card {{ border-left: 5px solid var(--red); padding: clamp(20px, 3vw, 34px); }}
    .data-card p {{ color: var(--muted); line-height: 1.7; max-width: 880px; margin: 0 0 22px; }}
    .data-actions {{ display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 28px; }}
    .data-actions a {{ display: inline-flex; align-items: center; min-height: 44px; padding: 0 16px; border: 1px solid var(--red); border-radius: 8px; color: var(--text); font-weight: 900; text-decoration: none; background: transparent; transition: background 220ms ease, transform 220ms ease, box-shadow 220ms ease; }}
    .data-actions a:hover {{ background: var(--red); transform: translateY(-2px); box-shadow: 0 16px 34px rgba(213,0,28,0.24); }}
    .data-table-wrap {{ overflow-x: auto; border: 1px solid var(--line); border-radius: 8px; }}
    .data-table {{ width: 100%; border-collapse: collapse; min-width: 760px; }}
    .data-table th {{ background: var(--red); color: var(--text); text-align: left; padding: 13px 14px; font-size: 13px; text-transform: uppercase; }}
    .data-table td {{ padding: 13px 14px; border-top: 1px solid var(--line); color: var(--muted); vertical-align: top; }}
    .data-table td:first-child {{ color: var(--text); font-weight: 800; }}
    .recommendations {{ display: grid; grid-template-columns: repeat(2, 1fr); gap: 16px; }}
    .recommendation {{ display: grid; grid-template-columns: 70px 1fr; gap: 18px; padding: 26px; position: relative; overflow: hidden; border-left: 5px solid var(--red); }}
    .rec-number {{ color: var(--red); font-size: 44px; line-height: 1; font-weight: 900; }}
    .recommendation h3 {{ margin: 0 0 12px; font-size: 21px; font-weight: 900; }}
    .recommendation p {{ color: var(--muted); font-style: italic; line-height: 1.62; margin-bottom: 18px; }}
    .impact-badge {{ display: inline-flex; align-items: center; min-height: 28px; padding: 0 10px; border-radius: 999px; color: #06120A; background: var(--green); font-size: 12px; font-weight: 900; text-transform: uppercase; margin-bottom: 10px; }}
    .recommendation strong {{ display: block; color: var(--text); line-height: 1.58; font-weight: 700; }}
    footer {{ border-top: 1px solid var(--line); padding: 30px 0 26px; color: var(--muted); font-size: 13px; text-align: center; }}
    .footer-car {{ display: block; margin: 0 auto 12px; color: var(--red); opacity: 0.6; }}
    @media (prefers-reduced-motion: reduce) {{
      section, .kpi-card, .data-actions a {{ transition: none; }}
    }}
    @media (max-width: 900px) {{
      .kpi-grid, .chart-grid, .recommendations {{ grid-template-columns: 1fr; }}
      header {{ min-height: 64vh; }}
      .section-title {{ display: block; }}
      .section-title p {{ margin-top: 10px; }}
      .recommendation {{ grid-template-columns: 1fr; }}
    }}
  </style>
</head>
<body>
  <header>
    <div class="wrap">
      <div class="brandline">{header_car}<div class="wordmark">PORSCHE</div></div>
      <h1>Global Performance Intelligence</h1>
      <p class="hero-copy">A synthetic, executive-grade BI dashboard built from Porsche's public delivery history and realistic model, region, revenue, satisfaction, dealer, and powertrain assumptions.</p>
    </div>
  </header>
  <nav>
    <div class="wrap">
      <a href="#executive">Executive Summary</a>
      <a href="#sales">Sales Trends</a>
      <a href="#regional">Regional</a>
      <a href="#profitability">Profitability</a>
      <a href="#electric">Electric Transition</a>
      <a href="#forecasting">Forecasting</a>
      <a href="#data-source">Data Source</a>
      <a href="#recommendations">Recommendations</a>
    </div>
  </nav>
  <main class="wrap">
    <section id="executive">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title">
        <h2>Executive Summary</h2>
        <p>Four board-level signals for 2024: scale, growth, portfolio concentration, and regional momentum.</p>
      </div>
      <div class="kpi-grid">{kpi_html}</div>
    </section>

    <section id="sales">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Sales Trend Analysis</h2><p>Model-line performance from 2018 through the 2024 transition year.</p></div>
      <div class="chart-card">{plot_html(trend)}</div>
      <details open><summary>Business Insight</summary><p>{insights["trend"]}</p></details>
    </section>

    <section id="regional">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Regional Breakdown</h2><p>Revenue contribution and delivery share show where Porsche's 2024 mix is strongest.</p></div>
      <div class="chart-grid">
        <div class="chart-card">{plot_html(bars)}</div>
        <div class="chart-card">{plot_html(donut)}</div>
      </div>
      <details open><summary>Business Insight</summary><p>{insights["regional"]}</p></details>
    </section>

    <section id="profitability">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Model Profitability Matrix</h2><p>Volume, transaction price, and revenue size plotted in one view.</p></div>
      <div class="chart-card">{plot_html(scatter)}</div>
      <details open><summary>Business Insight</summary><p>{insights["profit"]}</p></details>
    </section>

    <section id="electric">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Electric Transition Tracker</h2><p>Powertrain mix and revenue-per-unit economics across ICE, Hybrid, and Electric.</p></div>
      <div class="chart-card">{plot_html(transition)}</div>
      <details open><summary>Business Insight</summary><p>{insights["transition"]}</p></details>
    </section>

    <section id="forecasting">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Forecasting</h2><p>Simple linear regression baseline with conservative and optimistic EV adoption scenarios.</p></div>
      <div class="chart-card">{plot_html(forecast)}</div>
      <details open><summary>Business Insight</summary><p>{insights["forecast"]}</p></details>
    </section>

    <section id="data-source">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Data & Methodology</h2><p>Transparent source data, generation logic, and field definitions for auditability.</p></div>
      <div class="data-card">
        <p>This dashboard uses a synthetic dataset built from Porsche's publicly reported annual delivery figures for 2018-2024. Model, region, revenue, satisfaction, dealer, powertrain, and forecast values are generated with realistic assumptions so the dashboard can be shared publicly without private company data.</p>
        <div class="data-actions">
          <a href="{LIVE_DATA_URL}" target="_blank" rel="noreferrer">Download Excel Workbook</a>
          <a href="{SOURCE_URL}" target="_blank" rel="noreferrer">View Python Source</a>
        </div>
        <div class="data-table-wrap">
          <table class="data-table">
            <thead><tr><th>Column Name</th><th>Data Type</th><th>Description</th></tr></thead>
            <tbody>{dictionary_rows}</tbody>
          </table>
        </div>
      </div>
    </section>

    <section id="recommendations">
      <div class="section-watermark">{section_car}</div>
      <div class="section-title"><h2>Strategic Recommendations</h2><p>Consulting-style actions tied directly to the dashboard evidence.</p></div>
      <div class="recommendations">{recommendation_html}</div>
    </section>
  </main>
  <footer><div class="wrap">{footer_car}<div>Generated with Python, Pandas, Plotly, and HTML/CSS. Dataset is synthetic and for portfolio/business intelligence demonstration only.</div></div></footer>
  <script>
    const sections = Array.from(document.querySelectorAll("main section[id]"));
    const navLinks = Array.from(document.querySelectorAll("nav a"));

    const revealObserver = new IntersectionObserver((entries) => {{
      entries.forEach((entry) => {{
        if (entry.isIntersecting) entry.target.classList.add("visible");
      }});
    }}, {{ threshold: 0.16 }});

    sections.forEach((section) => revealObserver.observe(section));

    const activeObserver = new IntersectionObserver((entries) => {{
      const visible = entries
        .filter((entry) => entry.isIntersecting)
        .sort((a, b) => b.intersectionRatio - a.intersectionRatio)[0];
      if (!visible) return;
      const id = visible.target.id;
      navLinks.forEach((link) => link.classList.toggle("active", link.getAttribute("href") === `#${{id}}`));
    }}, {{ rootMargin: "-42% 0px -48% 0px", threshold: [0, 0.25, 0.5, 0.75, 1] }});

    sections.forEach((section) => activeObserver.observe(section));
    if (sections[0]) sections[0].classList.add("visible");

    const formatCount = (value, decimals, signed) => {{
      const fixed = Number(value).toFixed(decimals);
      const numeric = Number(fixed);
      const formatted = new Intl.NumberFormat("en-US", {{
        minimumFractionDigits: decimals,
        maximumFractionDigits: decimals
      }}).format(Math.abs(numeric));
      const sign = signed && numeric > 0 ? "+" : numeric < 0 ? "-" : "";
      return `${{sign}}${{formatted}}`;
    }};

    document.querySelectorAll("[data-count-target]").forEach((element) => {{
      const target = Number(element.dataset.countTarget);
      const prefix = element.dataset.countPrefix || "";
      const suffix = element.dataset.countSuffix || "";
      const decimals = Number(element.dataset.countDecimals || 0);
      const signed = element.dataset.countSigned === "true";
      const duration = 1300;
      const start = performance.now();
      const finalText = element.textContent;

      const tick = (now) => {{
        const progress = Math.min((now - start) / duration, 1);
        const eased = 1 - Math.pow(1 - progress, 3);
        const value = target * eased;
        element.textContent = `${{prefix}}${{formatCount(value, decimals, signed)}}${{suffix}}`;
        if (progress < 1) {{
          requestAnimationFrame(tick);
        }} else {{
          element.textContent = finalText;
        }}
      }};

      element.textContent = `${{prefix}}${{formatCount(0, decimals, signed)}}${{suffix}}`;
      requestAnimationFrame(tick);
    }});
  </script>
</body>
</html>"""


def main() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    df = build_dataset()
    df.to_csv(CSV_PATH, index=False)
    export_excel(df, EXCEL_PATH)
    HTML_PATH.write_text(make_dashboard(df), encoding="utf-8")
    print(f"Wrote {CSV_PATH}")
    print(f"Wrote {EXCEL_PATH}")
    print(f"Wrote {HTML_PATH}")


if __name__ == "__main__":
    main()
